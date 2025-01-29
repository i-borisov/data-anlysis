import pandas as pd
import openpyxl as op

df = pd.read_excel('/content/drive/MyDrive/tsum/TSUM - Задание python.xlsx', 'DATA')

brands = df['_brand'].unique()
brands.sort()

report = op.Workbook()
report.save('/content/drive/MyDrive/report.xlsx')

def select_brand(dataframe, brand_name):
  '''
  Функция выбирает из датафрейма строки с нужным брендом.
  '''
  brand_df = dataframe.loc[df['_brand'] == brand_name]
  return brand_df

def hiding_duplicates(column):
  '''
  Функция убирает повторяющиеся значения в столбцах финальной таблицы
  '''
  return column.where(column != column.shift())

def data_processing(brand_dataframe):
  '''
  Функция преобразует исходные данные в новый excel-файл.
  '''
  main_table = brand_dataframe.pivot_table(index=['_brand', '_div', '_rus1', '_rus2', '_rus3', '_rus4',], columns='_season', values='_stock_pcs', aggfunc='sum').reset_index()
  num_columns = main_table.shape[1]

  brand_values = main_table.iloc[:, [0] + list(range(6, num_columns))].groupby('_brand').agg('sum').reset_index()
  brand_values['Total'] = brand_values.sum(axis=1, numeric_only=True)
  div_values = main_table.iloc[:, [0, 1] + list(range(6, num_columns))].groupby(['_brand', '_div']).agg('sum').reset_index()
  div_values['Total'] = div_values.sum(axis=1, numeric_only=True)
  div_values.replace(0, '-', inplace=True)
  rus1_values = main_table.iloc[:, [0, 1, 2] + list(range(6, num_columns))].groupby(['_brand', '_div', '_rus1']).agg('sum').reset_index()
  rus1_values['Total'] = rus1_values.sum(axis=1, numeric_only=True)
  rus1_values.replace(0, '-', inplace=True)
  rus2_values = main_table.iloc[:, [0, 1, 2, 3] + list(range(6, num_columns))].groupby(['_brand', '_div', '_rus1', '_rus2']).agg('sum').reset_index()
  rus2_values['Total'] = rus2_values.sum(axis=1, numeric_only=True)
  rus2_values.replace(0, '-', inplace=True)
  rus3_values = main_table.iloc[:, [0, 1, 2, 3, 4] + list(range(6, num_columns))].groupby(['_brand', '_div', '_rus1', '_rus2', '_rus3']).agg('sum').reset_index()
  rus3_values['Total'] = rus3_values.sum(axis=1, numeric_only=True)
  rus3_values.replace(0, '-', inplace=True)
  rus4_values = main_table.iloc[:, [0, 1, 2, 3, 4, 5] + list(range(6, num_columns))].groupby(['_brand', '_div', '_rus1', '_rus2', '_rus3', '_rus4']).agg('sum').reset_index()
  rus4_values['Total'] = rus4_values.sum(axis=1, numeric_only=True)
  rus4_values.replace(0, '', inplace=True)

  concat_table = pd.concat([rus4_values, rus3_values, rus2_values, rus1_values, div_values, brand_values])

  result = concat_table.sort_values(['_div', '_rus1', '_rus2', '_rus3', '_rus4'], na_position='first')
  result[['_brand', '_div', '_rus1', '_rus2', '_rus3', '_rus4']] = result[['_brand', '_div', '_rus1', '_rus2', '_rus3', '_rus4']].apply(hiding_duplicates)
  result = result.fillna('')
  result['Total'] = result.sum(axis=1, numeric_only=True)

  excel_file = op.load_workbook('/content/drive/MyDrive/report.xlsx')
  excel_sheet = excel_file.create_sheet(title=brand)
  excel_data = result.values.tolist()

  for row in excel_data:
    excel_sheet.append(row)
  
  excel_file.save('/content/drive/MyDrive/report.xlsx')

for brand in brands:
  data_processing(select_brand(df, brand))
